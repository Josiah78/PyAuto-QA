import openpyxl

wb = openpyxl.Workbook()
wb.create_sheet('Sheet Number 2')
wb.get_sheet_names()
wb.remove_sheet(wb.get_sheet_by_name('Sheet Number 2'))
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')
sheetsheet['A1']
sheet['A1'].value
shee['A1'] = 'Apples'
sheet['A1'].value

for column in range(0,3):
    sheet.cell(row=1, column=column+1).value = 'Cheese'
wb.sabe('cheese.xlsx')