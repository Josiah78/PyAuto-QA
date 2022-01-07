import openpyxl

wb = openpyxl.load_workbook('Students.xlsx')
wb
wb.get_sheet_names()
wb.close()
wb = openpyxl.load_workbook('Students.xlsx')
sheet = wb.get_sheet_by_name('Students')
sheet
sheet = wb.active
sheet
sheet.title
sheet['A1']
sheet['A1'].value
sheet['B4'].value
cell = sheet['B4']
cell.row
cell.value
cell.coordinate
sheet.cell(row=1, column=1)
sheet.cell(row=1, column=1).value
for i in range(0,3):
    print(sheet.cell(row=1, column=i+1).value)

sheet['A1':'C5']
sheet.columns
sheet.rows
list(sheet.columns)[1]
list(sheet.rows)[1]
for row in sheet['A1':'C5']:
    for cell in row:
        print(cell.value)
    print('\n')

for name in list(sheet.columns)[1]:
    print(name.value)